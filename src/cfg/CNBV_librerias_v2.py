# ----------------------------------------------------------------------------------------
#                                  LIBRERIAS NECESARIAS 
# Descripción: Abajo listo las librerías necesarias para su ejecución x pasos
# Autor: SteveCarpio-2024
# ----------------------------------------------------------------------------------------
# Librerias invocadas en los PASOS del programa :                 # Ma P0 P1 P2 P3 P4 P5 #
import os                                                         # Ma p0 -- p2 -- p4 p5 #
import sys                                                        # Ma -- -- -- -- p4    #
import glob                                                       # -- p0 -- p2 -- p4    #
import time                                                       # -- -- p1             #
import pandas as pd                                               # -- -- -- p2 p3 p4 p5 #
import subprocess                                                 # -- -- -- -- p3       #
import openpyxl                                                   # -- -- -- -- -- -- p5 #
from selenium import webdriver                                    # -- -- p1             #
from selenium.webdriver.chrome.service import Service             # -- -- p1             #
from selenium.webdriver.common.by import By                       # -- -- p1             #
from bs4 import BeautifulSoup                                     # -- -- -- p2          #
from openpyxl import Workbook                                     # -- -- -- -- -- -- p5 #
from openpyxl import load_workbook                                # -- -- -- -- -- -- p5 #
from openpyxl.styles import Font, PatternFill                     # -- -- -- -- -- -- p5 #
from openpyxl.utils import get_column_letter                      # -- -- -- -- -- -- p5 #
from datetime import datetime as dt                               # -- -- -- -- -- -- p5 #
# ----------------------------------------------------------------------------------------
