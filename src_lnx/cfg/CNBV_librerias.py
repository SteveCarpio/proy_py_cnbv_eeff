# -------------------------------------------------------------------------------------------
#                                  LIBRERIAS NECESARIAS 
# Descripción: Abajo listo las librerías necesarias para su ejecución x pasos
# Autor: SteveCarpio-2025
# -------------------------------------------------------------------------------------------
# Librerias invocadas en los PASOS del programa :                 # Ma P0 P1 P2 P3 P4 P5 P6 #
import os                                                         # Ma p0 -- p2 -- p4 p5    #
import sys                                                        # Ma -- -- -- -- p4       #
import glob                                                       # -- p0 -- p2 -- p4       #
import time                                                       # -- -- p1                #
import pandas as pd                                               # -- -- -- p2 p3 p4 p5    #
import subprocess                                                 # -- -- -- -- p3          #
import openpyxl                                                   # -- -- -- -- -- -- p5    #
import smtplib                                                    # -- -- -- -- -- -- -- p6 #
import re                                                         # -- -- -- -- -- -- -- p6 #
from colorama import init, Fore, Back, Style                      # Ma                      #
from selenium import webdriver                                    # -- -- p1 -- -- -- -- -- #
from selenium.webdriver.chrome.service import Service             # -- -- p1 -- -- -- -- -- #
from selenium.webdriver.chrome.options import Options             # -- -- p1 -- -- -- -- -- #
from selenium.webdriver.support.ui import WebDriverWait           # -- -- p1 -- -- -- -- -- #
from selenium.webdriver.support import expected_conditions as EC  # -- -- p1 -- -- -- -- -- #
from selenium.webdriver.common.by import By                       # -- -- p1 -- -- -- -- -- #
from bs4 import BeautifulSoup                                     # -- -- -- p2             #
from openpyxl import Workbook                                     # -- -- -- -- -- -- p5    #
from openpyxl import load_workbook                                # -- -- -- -- -- -- p5    #
from openpyxl.styles import Font, PatternFill                     # -- -- -- -- -- -- p5    #
from openpyxl.utils import get_column_letter                      # -- -- -- -- -- -- p5    #
from datetime import datetime as dt                               # -- -- -- -- -- -- p5    #
from email.mime.multipart import MIMEMultipart                    # -- -- -- -- -- -- -- p6 #
from email.mime.text import MIMEText                              # -- -- -- -- -- -- -- p6 #
from email.mime.application import MIMEApplication                # -- -- -- -- -- -- -- p6 #
# -------------------------------------------------------------------------------------------

# CUSTOMIZE LAS LIBRERÍAS ---------------------------------------------------------------------------------------------------
chrome_options = Options()
prefs = {"profile.managed_default_content_settings.images": 2 ,       # 2 = Bloquear  Imágenes
         "profile.managed_default_content_settings.javascript": 1}    # 1 = Habilitar JavaScript 
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--disable-extensions")          # Desactivar extensiones
chrome_options.add_argument("--log-level=3")                 # Suprime los mensajes de log (nivel de error: WARNING y superior)
chrome_options.add_argument("--disable-sync")                # Desactivar la sincronización con el perfil de usuario
chrome_options.add_argument("--disable-blink-features=AutomationControlled") # Desactivar animaciones
chrome_options.add_argument("--disable-features=VizDisplayCompositor")       # Desactivar animaciones
chrome_options.add_argument("--disable-plugins")             # Desactivar plugins
chrome_options.add_argument("--incognito")                   # Usar modo incógnito
chrome_options.add_argument("--headless")                    # Ejecutar SIN interfaz gráfica :  carga más rápido
chrome_options.add_argument("--disable-gpu")                 # Desactivar uso de GPU         :  carga más rápido
chrome_options.add_argument('--ignore-certificate-errors')   # Ignorar certificados   
#chrome_options.add_argument("--user-data-dir=/path/to/a/clean/profile")   # Con esta option me hace cosas raras
# ----------------------------------------------------------------------------------------------------------------------------